"""Seed production-grade demo knowledge assets into the relational index."""

from __future__ import annotations

import io
import textwrap
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select

from app.core.db import db_session
from app.models.relational import KnowledgeChunk, KnowledgeDocument, KnowledgeObjectLink
from app.services.ai.knowledge_ingestion import (
    CHUNKS,
    DOCUMENTS,
    JOBS,
    markdown_to_chunks,
    parse_to_markdown_with_metadata,
)


ROOT_DIR = Path(__file__).resolve().parents[4]
ASSET_DIR = ROOT_DIR / "data" / "knowledge_assets"


RISK_ROWS = [
    ("MB-7781", "北辰电子材料", "锡膏 S12", "2026-05-19 08:42", "运输温控记录缺失 47 分钟", "WO-260521-017;WO-260521-021", "high", "SQE-刘洋", "2026-05-22", "冻结批次，补交温度曲线和开封回温记录"),
    ("MB-7781", "北辰电子材料", "锡膏 S12", "2026-05-19 08:42", "二次回温记录未签名", "WO-260521-017", "medium", "QE-王敏", "2026-05-22", "复核领用记录，增加首件 X-Ray"),
    ("MB-7781", "北辰电子材料", "锡膏 S12", "2026-05-19 08:42", "供应商 D3 围堵照片缺失", "WO-260521-017", "medium", "SQE-刘洋", "2026-05-22", "补充供应商库存冻结证据"),
    ("MB-7784", "北辰电子材料", "锡膏 S12", "2026-05-21 10:05", "运输箱温度记录器电量低", "WO-260522-006", "medium", "仓储-冯宇", "2026-05-24", "换用新记录器并抽检黏度"),
    ("MB-7785", "北辰电子材料", "锡膏 S12", "2026-05-21 15:30", "COA 批号与外箱标签不一致", "WO-260522-009", "high", "SQE-刘洋", "2026-05-24", "暂停入库，供应商重发 COA"),
    ("MB-7790", "华东焊材", "锡膏 S18", "2026-05-22 09:10", "开封后使用时长超过 8 小时", "WO-260522-012", "medium", "PE-黄婷", "2026-05-25", "退回余料并复核工艺纪律"),
    ("FL-2431", "安特化学", "助焊剂 F8", "2026-05-20 11:20", "供应商等级下降", "WO-260521-030", "medium", "IQC-赵宁", "2026-05-25", "提高来料抽检比例至 30%"),
    ("PK-5530", "森和包装", "防潮袋", "2026-05-18 14:00", "湿度卡变色", "WO-260520-014", "high", "仓储-冯宇", "2026-05-21", "隔离包装批次，复核 MSL 管控"),
    ("MB-7782", "北辰电子材料", "锡膏 S12", "2026-05-20 09:15", "记录完整，未发现异常", "-", "low", "仓储-冯宇", "2026-05-23", "维持正常来料检验"),
    ("MB-7783", "北辰电子材料", "锡膏 S12", "2026-05-20 16:40", "供应商已补齐温控记录", "WO-260521-040", "low", "SQE-刘洋", "2026-05-23", "保留抽检，观察两批"),
    ("CN-9012", "科源电子", "BGA 控制芯片", "2026-05-21 13:30", "MSL 标签缺少开封时间", "WO-260522-015", "medium", "IQC-赵宁", "2026-05-24", "烘烤后复检，补录标签"),
    ("TR-3308", "长江电子", "连接器", "2026-05-22 08:20", "端子氧化投诉", "SO-8821", "high", "SQE-刘洋", "2026-05-26", "启动供应商 8D 和库存复核"),
]


TEMP_PROFILE_ROWS = [
    ("09:00", 238.1, 239.5, 241.2, "normal"),
    ("09:05", 237.9, 238.6, 240.8, "normal"),
    ("09:10", 236.4, 235.7, 239.9, "warning"),
    ("09:12", 235.1, 233.8, 239.2, "warning"),
    ("09:15", 233.7, 232.6, 238.1, "warning"),
    ("09:20", 232.9, 232.2, 237.5, "warning"),
    ("09:25", 233.5, 232.9, 237.8, "warning"),
    ("09:30", 235.6, 234.7, 238.7, "warning"),
    ("09:35", 238.3, 237.5, 240.4, "recovered"),
    ("09:40", 239.1, 238.8, 241.0, "normal"),
    ("09:45", 239.0, 239.1, 241.3, "normal"),
    ("09:50", 238.8, 239.0, 241.2, "normal"),
]


DEMO_ASSETS: list[dict[str, Any]] = [
    {
        "document_id": "kb-doc-quality-sop-docx",
        "file_name": "SOP-QA-014-solder-void-recheck.docx",
        "title": "SOP-QA-014 焊点虚焊复检与批次冻结流程",
        "kind": "docx",
        "permission_scope": "department",
        "owner_user_id": "qe_wang",
        "blocks": [
            ("h1", "SOP-QA-014 焊点虚焊复检与批次冻结流程"),
            ("p", "文件编号：SOP-QA-014；版本：V3.3；生效日期：2026-05-27；适用工厂：上海一厂 SMT 车间；文件所有者：质量管理部。"),
            ("table", [["字段", "内容"], ["适用产品", "电控模块、传感控制板、高压连接控制板"], ["适用产线", "SMT-01 至 SMT-06"], ["关键对象", "QE-20260521-001 / WO-260521-017 / MB-7781 / SMT-03"], ["保存年限", "3 年"]]),
            ("h2", "1. 目的和范围"),
            ("p", "本文件规定 BGA 焊点虚焊、空焊、少锡等缺陷被 AOI、X-Ray 或客户退货连续检出后的隔离、复检、根因确认、CAPA 立项、放行审批和记录归档要求。"),
            ("h2", "2. 启动条件"),
            ("table", [["触发条件", "阈值", "处理等级"], ["BGA 虚焊率", "连续 2 小时超过 2.0%", "重大质量异常"], ["同批次重复缺陷", "同一锡膏批次涉及 3 个以上工单", "批次冻结"], ["回流焊温区偏移", "温区 5 或 6 偏离中心线超过 5 C", "设备复核"], ["供应商证据缺失", "冷链、COA、开封回温记录任一缺失", "供应商 8D"]]),
            ("h2", "3. RACI 职责"),
            ("table", [["角色", "职责", "输出记录"], ["质量工程师", "创建质量事件、组织 X-Ray 复检、维护证据链", "复检报告、缺陷判定"], ["生产主管", "冻结 WIP、暂停未经确认放行、组织首件复核", "冻结清单、首件记录"], ["设备工程师", "调取炉温曲线、点检 TEMP-05、确认报警阈值", "设备复核报告"], ["SQE", "向供应商索取 8D、冷链温度和批次 COA", "8D 跟踪台账"], ["质量经理", "审批 CAPA 和最终放行", "放行审批"]]),
            ("h2", "4. 判定标准"),
            ("p", "AOI 判定为疑似虚焊时，必须由 X-Ray 或切片复核确认。BGA 焊点空洞面积超过 IPC-A-610 Class 2 内控阈值 25% 时判定为不合格；同一器件相邻焊点连续异常时按重大缺陷处理。若 AOI 图片、X-Ray 图片、炉温曲线任一证据缺失，不得关闭质量事件。"),
            ("h2", "5. 标准处置流程"),
            ("table", [["步骤", "操作", "时限", "系统记录"], ["1", "MES 冻结同工单、同班次、同锡膏批次和同回流焊窗口产品", "30 分钟内", "HOLD-QA"], ["2", "仓库锁定 MB-7781 剩余库存并粘贴红色 HOLD 标签", "1 小时内", "库存状态变更"], ["3", "质量工程师对前后各 2 小时产品进行 X-Ray 100% 复检", "4 小时内", "X-Ray 复检报告"], ["4", "设备工程师复核 SMT-03 温区 5、6 曲线和 TEMP-05 校准", "4 小时内", "MNT-RPT"], ["5", "SQE 核对供应商 8D、冷链温度和开封回温记录", "24 小时内", "8D 台账"], ["6", "质量经理基于证据决定返工、报废、让步或放行", "48 小时内", "放行审批"]]),
            ("h2", "6. 放行门禁"),
            ("p", "受影响批次只有在复检合格、供应商证据补齐、设备点检无异常、CAPA 临时围堵完成、质量经理审批完成后才能释放。影响客户订单 SO-8821 时，销售对外交付承诺必须引用质量批准后的替代批次。"),
            ("h2", "7. 记录清单"),
            ("table", [["记录", "责任人", "必填字段"], ["AOI 缺陷截图", "质量工程师", "缺陷代码、坐标、图片编号"], ["X-Ray 复检记录", "质量工程师", "工单、序列号、判定结果"], ["炉温曲线", "设备工程师", "配方、温区、时间窗口"], ["供应商 8D", "SQE", "D1-D8 状态、根因、验证结果"], ["CAPA", "质量经理", "围堵、根因、永久措施、有效性"]]),
        ],
        "links": [("QualityEvent", "QE-20260521-001", "AOI 焊点虚焊异常"), ("WorkOrder", "WO-260521-017", "电控模块工单"), ("Equipment", "SMT-03", "SMT-03 回流焊"), ("MaterialBatch", "MB-7781", "锡膏 S12 批次 MB-7781")],
    },
    {
        "document_id": "kb-doc-capa-072-docx",
        "file_name": "CAPA-072-solder-void-closure-report.docx",
        "title": "CAPA-072 焊点虚焊闭环报告",
        "kind": "docx",
        "permission_scope": "department",
        "owner_user_id": "qe_wang",
        "blocks": [
            ("h1", "CAPA-072 焊点虚焊闭环报告"),
            ("p", "CAPA 编号：CAPA-072；来源事件：QE-20260521-001；负责人：王敏；状态：有效性验证中；目标关闭日期：2026-06-15。"),
            ("table", [["阶段", "结论", "证据"], ["问题描述", "SMT-03 产线 BGA 焊点虚焊率升高至 2.8%", "AOI/X-Ray 报告"], ["临时围堵", "冻结 WO-260521-017 及前后窗口产品 1,248 件", "MES HOLD-QA 清单"], ["根因", "锡膏 MB-7781 冷链记录缺失叠加回流焊温区 5 波动", "供应商 8D、设备报告"], ["永久措施", "增加温度记录器校验、TEMP-05 接头更换、首件 X-Ray", "行动计划"], ["有效性验证", "连续 5 批 BGA 虚焊率低于 0.6%", "SPC 趋势"]]),
            ("h2", "5 Why 根因分析"),
            ("table", [["Why", "问题", "回答"], ["1", "为什么出现虚焊？", "焊点润湿不足，X-Ray 显示空洞和少锡。"], ["2", "为什么润湿不足？", "锡膏活性下降且回流峰值偏低。"], ["3", "为什么锡膏活性下降？", "MB-7781 运输温控记录缺失，回温记录不完整。"], ["4", "为什么峰值偏低？", "SMT-03 温区 5 接头阻值偏高导致控制偏移。"], ["5", "为什么未提前发现？", "周点检只看停机报警，未纳入趋势偏移预警。"]]),
            ("h2", "行动计划"),
            ("table", [["行动", "负责人", "截止日期", "验收标准"], ["供应商温控记录器出库扫码", "SQE-刘洋", "2026-06-01", "100% 批次可追溯"], ["TEMP-05 接头更换与端子紧固", "设备-周强", "2026-05-29", "连续 7 天无温区漂移"], ["首件 X-Ray 检验纳入作业指导书", "QE-王敏", "2026-06-03", "执行率 100%"], ["MES 冻结原因标准化", "数据专员-何静", "2026-06-05", "HOLD-QA 原因码上线"]]),
        ],
        "links": [("CAPA", "CAPA-072", "焊点虚焊 CAPA"), ("QualityEvent", "QE-20260521-001", "AOI 焊点虚焊异常"), ("Equipment", "SMT-03", "SMT-03 回流焊")],
    },
    {
        "document_id": "kb-doc-supplier-8d-xlsx",
        "file_name": "supplier-8d-MB-7781-risk-ledger.xlsx",
        "title": "北辰电子材料 MB-7781 供应商 8D 风险台账",
        "kind": "xlsx",
        "permission_scope": "team",
        "owner_user_id": "scm_liu",
        "sheets": {
            "RiskLedger": [["batch", "supplier", "material", "received_at", "issue", "affected_orders", "risk", "owner", "due", "action"], *RISK_ROWS],
            "TemperatureLog": [["time", "transport_c", "warehouse_c", "return_warm_min", "status"], *[(f"2026-05-19 {hour:02d}:00", 2.5 + (hour % 5) * 0.8, 21.5 + (hour % 3), 45 + hour * 3, "gap" if hour in (7, 8) else "ok") for hour in range(24)]],
            "InspectionRule": [["material", "condition", "incoming_ratio", "test_method", "release_rule"], ["锡膏 S12", "缺少冷链记录或 AOI 虚焊率超过 2.0%", "100%", "COA 核对 + 黏度测试 + 首件 X-Ray", "SQE 与质量经理双签后释放"], ["锡膏 S12", "供应商 8D 未关闭", "100%", "每批留样 + 回温记录复核", "CAPA 有效性验证通过后恢复抽检"], ["助焊剂 F8", "供应商等级下降或批次投诉", "30%", "外观 + 固含量抽检", "来料检验合格后释放"], ["BGA 控制芯片", "MSL 标签缺失", "100%", "烘烤记录 + 外观 + 首件", "PE 与 QE 双签"], ["连接器", "客户投诉或端子氧化", "50%", "外观 + 插拔力", "SQE 关闭 8D 后恢复"]],
            "EightDStatus": [["step", "status", "evidence", "gap"], ["D1 Team", "completed", "供应商、SQE、质量工程师已确认联系人", "-"], ["D2 Problem", "completed", "AOI 和 X-Ray 均指向 BGA 虚焊", "-"], ["D3 Containment", "open", "MB-7781 已冻结", "供应商库存冻结照片未回传"], ["D4 Root Cause", "open", "疑似冷链温控中断和回温时间不足", "缺少运输段原始温度曲线"], ["D5 Corrective Action", "planned", "增加温度记录器和出库扫码校验", "有效性验证未完成"], ["D6 Validation", "planned", "连续 5 批缺陷率低于 0.6%", "等待生产验证"], ["D7 Prevention", "draft", "供应商稽核清单更新", "未发布"], ["D8 Closure", "not_started", "-", "等待 D6/D7"]],
            "ActionTracker": [["id", "action", "owner", "due", "status", "evidence"], *[(f"A-{i:03d}", action, owner, due, status, evidence) for i, (action, owner, due, status, evidence) in enumerate([("补交 MB-7781 原始温度曲线", "北辰电子材料", "2026-05-22", "open", "邮件追踪"), ("冻结供应商在途同批材料", "SQE-刘洋", "2026-05-22", "done", "冻结照片"), ("复核仓储回温 SOP", "仓储-冯宇", "2026-05-23", "done", "培训签到"), ("首件 X-Ray 记录模板更新", "QE-王敏", "2026-05-24", "open", "模板草案"), ("供应商出库扫码改造", "北辰电子材料", "2026-06-01", "planned", "项目计划"), ("连续 5 批验证", "QE-王敏", "2026-06-12", "planned", "SPC 报表")], start=1)]],
        },
        "links": [("Supplier", "SUP-BEICHEN", "北辰电子材料"), ("MaterialBatch", "MB-7781", "锡膏 S12 批次 MB-7781"), ("CAPA", "CAPA-072", "焊点虚焊 CAPA")],
    },
    {
        "document_id": "kb-doc-process-control-xlsx",
        "file_name": "SMT-BGA-process-control-plan.xlsx",
        "title": "SMT BGA 工艺控制计划",
        "kind": "xlsx",
        "permission_scope": "enterprise",
        "owner_user_id": "pe_huang",
        "sheets": {
            "ControlPlan": [["process_step", "parameter", "target", "lsl", "usl", "frequency", "reaction_plan"], ["锡膏回温", "回温时间", "4h", "3.5h", "6h", "每批", "不足时禁止上线"], ["锡膏印刷", "钢网张力", "38N/cm", "35", "45", "每班", "停线清洁钢网"], ["贴片", "BGA 偏移", "0.00mm", "-0.08", "0.08", "每小时", "校准贴片机"], ["回流", "温区 5", "239C", "236", "244", "每 5 分钟", "设备复核"], ["回流", "峰值温度", "245C", "241", "249", "每炉", "运行黄金板"], ["X-Ray", "空洞率", "<25%", "0", "25", "首件/异常", "冻结批次"]],
            "SPCTrend": [["timestamp", "line", "parameter", "value", "ucl", "lcl", "status"], *[(f"2026-05-21 {8 + i // 4:02d}:{(i % 4) * 15:02d}", "SMT-03", "Zone5", 239 - (6 if 5 <= i <= 8 else 0) + (i % 3) * 0.4, 244, 236, "alert" if 5 <= i <= 8 else "normal") for i in range(24)]],
            "ReactionMatrix": [["condition", "severity", "containment", "owner"], ["虚焊率 > 2.0%", "major", "冻结同班次 WIP", "QE"], ["温区偏移 > 5C", "major", "运行黄金板并锁定设备", "Maintenance"], ["冷链记录缺失", "major", "冻结供应商批次", "SQE"], ["客户订单受影响", "critical", "销售暂停承诺并升级质量经理", "Quality Manager"]],
        },
        "links": [("Equipment", "SMT-03", "SMT-03 回流焊"), ("Sensor", "TEMP-05", "温区 5 温度传感器"), ("ProcessParameter", "ZONE5", "回流焊温区 5")],
    },
    {
        "document_id": "kb-doc-maintenance-log-pdf",
        "file_name": "SMT-03-zone5-temperature-review.pdf",
        "title": "SMT-03 回流焊温区 5 波动复核报告",
        "kind": "pdf",
        "permission_scope": "enterprise",
        "owner_user_id": "mm_zhou",
        "pages": [
            ["Maintenance Review Report: SMT-03 Reflow Oven Zone 5", "Report No.: MNT-RPT-20260521-03", "Owner: Maintenance Dept. Reviewer: Zhou Qiang", "Event link: QE-20260521-001, Work order WO-260521-017, Sensor TEMP-05", "Time window: 2026-05-21 09:12 to 09:35. Recipe: EC-CTRL-BGA-V2. Line: SMT-03."],
            ["Measured Temperature Profile", "09:10 Zone5 actual 235.7 C vs target 239.0 C", "09:15 Zone5 actual 232.6 C vs target 239.0 C", "09:20 Zone5 actual 232.2 C vs target 239.0 C", "09:35 Zone5 recovered to 237.5 C", "No PLC hard alarm was triggered because the stop threshold was not reached."],
            ["Engineering Assessment", "Finding 1: Zone 5 actual temperature stayed 6.8 C below recipe center for 11 minutes.", "Finding 2: TEMP-05 calibration is valid until 2026-06-30, but connector resistance was above the internal warning limit.", "Impact: The drift may increase solder void risk when solder paste warm-up time is short or cold-chain evidence is missing.", "Immediate action: Run golden-board profile verification before restart and hold products processed from 09:00 to 10:00.", "Corrective action: Replace TEMP-05 connector, tighten terminal block, and add weekly trend review for zones 4 to 6.", "Release condition: Maintenance sign-off, quality X-Ray recheck, and production first-piece approval must be completed."],
        ],
        "links": [("Equipment", "SMT-03", "SMT-03 回流焊"), ("Sensor", "TEMP-05", "温区 5 温度传感器"), ("WorkOrder", "WO-260521-017", "电控模块工单")],
    },
    {
        "document_id": "kb-doc-customer-risk-pdf",
        "file_name": "SO-8821-customer-risk-communication-standard.pdf",
        "title": "SO-8821 客户交付风险沟通标准",
        "kind": "pdf",
        "permission_scope": "enterprise",
        "owner_user_id": "pm_li",
        "pages": [
            ["Customer Delivery Risk Communication Standard", "Document: STD-SALES-QA-009", "Applies to customer order SO-8821 and all quality events that affect confirmed delivery.", "The sales team must not promise a recovery date before quality confirms isolation scope and substitute batch availability."],
            ["Required inputs before external communication", "1. Quality event number and affected serial range.", "2. Frozen quantity, released quantity, and substitute batch quantity.", "3. CAPA status and temporary containment result.", "4. Production recovery plan approved by the production manager.", "5. Customer-facing risk level approved by the quality manager."],
            ["Escalation and message template", "Critical: customer shipment is blocked or regulatory impact is possible.", "Major: delivery date may move by more than 24 hours.", "Standard wording: affected products are under internal containment; confirmed conforming inventory will be allocated first; updated delivery commitment will be provided after quality release."],
        ],
        "links": [("CustomerOrder", "SO-8821", "客户订单 SO-8821"), ("QualityEvent", "QE-20260521-001", "AOI 焊点虚焊异常"), ("CAPA", "CAPA-072", "焊点虚焊 CAPA")],
    },
]


def _w_text(text: str, *, bold: bool = False, size: int = 22) -> str:
    bold_xml = "<w:b/>" if bold else ""
    return f'<w:r><w:rPr>{bold_xml}<w:sz w:val="{size}"/></w:rPr><w:t xml:space="preserve">{escape(text)}</w:t></w:r>'


def _w_paragraph(text: str, *, style: str | None = None) -> str:
    style_xml = f'<w:pPr><w:pStyle w:val="{style}"/></w:pPr>' if style else ""
    bold = style in {"Title", "Heading1", "Heading2"}
    size = 32 if style == "Title" else 26 if style == "Heading1" else 24 if style == "Heading2" else 22
    return f"<w:p>{style_xml}{_w_text(text, bold=bold, size=size)}</w:p>"


def _w_table(rows: list[list[str]]) -> str:
    tr_xml = []
    for row_index, row in enumerate(rows):
        cells = []
        for cell in row:
            shade = '<w:shd w:fill="D9EAF7"/>' if row_index == 0 else ""
            cells.append(
                "<w:tc><w:tcPr>"
                '<w:tcW w:w="2400" w:type="dxa"/>'
                f"{shade}</w:tcPr>{_w_paragraph(str(cell))}</w:tc>"
            )
        tr_xml.append("<w:tr>" + "".join(cells) + "</w:tr>")
    return (
        "<w:tbl><w:tblPr><w:tblStyle w:val=\"TableGrid\"/>"
        '<w:tblBorders><w:top w:val="single" w:sz="4"/><w:left w:val="single" w:sz="4"/>'
        '<w:bottom w:val="single" w:sz="4"/><w:right w:val="single" w:sz="4"/>'
        '<w:insideH w:val="single" w:sz="4"/><w:insideV w:val="single" w:sz="4"/></w:tblBorders>'
        "</w:tblPr>" + "".join(tr_xml) + "</w:tbl>"
    )


def _docx_bytes(blocks: list[tuple[str, Any]]) -> bytes:
    body_parts = []
    for kind, payload in blocks:
        if kind == "h1":
            body_parts.append(_w_paragraph(str(payload), style="Title"))
        elif kind == "h2":
            body_parts.append(_w_paragraph(str(payload), style="Heading1"))
        elif kind == "table":
            body_parts.append(_w_table(payload))
        else:
            body_parts.append(_w_paragraph(str(payload)))

    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{''.join(body_parts)}<w:sectPr /></w:body></w:document>"
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:style w:type="paragraph" w:styleId="Title"><w:name w:val="Title"/></w:style>'
        '<w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="heading 1"/></w:style>'
        '<w:style w:type="table" w:styleId="TableGrid"><w:name w:val="Table Grid"/></w:style>'
        "</w:styles>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        "</Relationships>"
    )
    document_rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        "</Relationships>"
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", rels)
        archive.writestr("word/_rels/document.xml.rels", document_rels)
        archive.writestr("word/document.xml", document_xml)
        archive.writestr("word/styles.xml", styles_xml)
    return buffer.getvalue()


def _xlsx_bytes(sheets: dict[str, list[list[Any]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name[:31])
        for row in rows:
            sheet.append(list(row))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 42)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _pdf_bytes(pages: list[list[str]]) -> bytes:
    try:
        import pymupdf  # type: ignore[import-not-found]
    except Exception as exc:  # pragma: no cover
        try:
            import fitz as pymupdf  # type: ignore[no-redef]
        except Exception as fitz_exc:  # pragma: no cover
            return _simple_pdf_bytes(pages)

    document = pymupdf.open()
    for page_lines in pages:
        page = document.new_page(width=595, height=842)
        y = 56
        for index, line in enumerate(page_lines):
            font_size = 15 if index == 0 else 10.5
            wrapped = textwrap.wrap(str(line), width=88) or [""]
            for wrapped_line in wrapped:
                page.insert_text((56, y), wrapped_line, fontsize=font_size, fontname="helv")
                y += 18 if index == 0 else 14
            y += 4
    return document.tobytes()


def _escape_pdf_text(text: str) -> str:
    return str(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _simple_pdf_bytes(pages: list[list[str]]) -> bytes:
    """Create a tiny text PDF when PyMuPDF is unavailable in lightweight tests."""

    objects: list[bytes] = []
    page_object_ids: list[int] = []
    font_object_id = 3
    next_object_id = 4

    for page_lines in pages:
        content_object_id = next_object_id
        page_object_id = next_object_id + 1
        next_object_id += 2
        page_object_ids.append(page_object_id)

        commands = ["BT", "/F1 11 Tf", "56 790 Td", "14 TL"]
        for index, line in enumerate(page_lines):
            font_size = 15 if index == 0 else 10.5
            commands.append(f"/F1 {font_size} Tf")
            wrapped = textwrap.wrap(str(line), width=88) or [""]
            for wrapped_line in wrapped:
                commands.append(f"({_escape_pdf_text(wrapped_line)}) Tj")
                commands.append("T*")
            commands.append("T*")
        commands.append("ET")
        stream = "\n".join(commands).encode("latin-1", errors="replace")
        objects.append(
            f"{content_object_id} 0 obj\n<< /Length {len(stream)} >>\nstream\n".encode("ascii")
            + stream
            + b"\nendstream\nendobj\n"
        )
        objects.append(
            f"{page_object_id} 0 obj\n"
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 {font_object_id} 0 R >> >> "
            f"/Contents {content_object_id} 0 R >>\n"
            f"endobj\n"
        .encode("ascii"))

    kids = " ".join(f"{page_id} 0 R" for page_id in page_object_ids)
    header_objects = [
        b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n",
        f"2 0 obj\n<< /Type /Pages /Kids [{kids}] /Count {len(page_object_ids)} >>\nendobj\n".encode("ascii"),
        b"3 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\nendobj\n",
    ]
    all_objects = header_objects + objects
    pdf = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for obj in all_objects:
        offsets.append(len(pdf))
        pdf.extend(obj)
    xref_offset = len(pdf)
    total_objects = len(all_objects) + 1
    pdf.extend(f"xref\n0 {total_objects}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {total_objects} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    return bytes(pdf)


def _asset_bytes(asset: dict[str, Any]) -> bytes:
    if asset["kind"] == "docx":
        return _docx_bytes(asset["blocks"])
    if asset["kind"] == "xlsx":
        return _xlsx_bytes(asset["sheets"])
    if asset["kind"] == "pdf":
        return _pdf_bytes(asset["pages"])
    raise ValueError(f"Unsupported demo asset kind: {asset['kind']}")


async def seed_demo_knowledge_assets() -> dict[str, int]:
    tenant_id = 1
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    seeded_documents = 0
    seeded_chunks = 0

    async with db_session() as session:
        for asset in DEMO_ASSETS:
            source_path = ASSET_DIR / asset["file_name"]
            content = _asset_bytes(asset)
            source_path.write_bytes(content)
            source_type, markdown, metadata = parse_to_markdown_with_metadata(asset["file_name"], content)
            updated_at = datetime.now()
            document = {
                "document_id": asset["document_id"],
                "tenant_id": tenant_id,
                "source_file_name": asset["file_name"],
                "source_type": source_type,
                "title": asset["title"],
                "markdown_content": markdown,
                "ocr_result": metadata.get("ocr_result"),
                "permission_scope": asset["permission_scope"],
                "owner_user_id": asset["owner_user_id"],
                "source_path": str(source_path),
                "status": "indexed",
                "updated_at": updated_at.isoformat(),
            }
            chunks = markdown_to_chunks(markdown, asset["document_id"], asset["permission_scope"], tenant_id=tenant_id)
            for index, chunk in enumerate(chunks, start=1):
                chunk["chunk_id"] = f"{asset['document_id']}-chunk-{index:02d}"

            row = await session.scalar(
                select(KnowledgeDocument).where(
                    KnowledgeDocument.tenant_id == tenant_id,
                    KnowledgeDocument.document_id == asset["document_id"],
                )
            )
            if row:
                row.source_file_name = document["source_file_name"]
                row.source_type = document["source_type"]
                row.title = document["title"]
                row.markdown_content = document["markdown_content"]
                row.ocr_result = document["ocr_result"]
                row.permission_scope = document["permission_scope"]
                row.owner_user_id = document["owner_user_id"]
                row.source_path = document["source_path"]
                row.status = document["status"]
                row.updated_at = updated_at
            else:
                session.add(KnowledgeDocument(
                    tenant_id=tenant_id,
                    document_id=document["document_id"],
                    source_file_name=document["source_file_name"],
                    source_type=document["source_type"],
                    title=document["title"],
                    markdown_content=document["markdown_content"],
                    ocr_result=document["ocr_result"],
                    permission_scope=document["permission_scope"],
                    owner_user_id=document["owner_user_id"],
                    source_path=document["source_path"],
                    status=document["status"],
                    updated_at=updated_at,
                ))
                seeded_documents += 1

            await session.execute(
                delete(KnowledgeChunk).where(
                    KnowledgeChunk.tenant_id == tenant_id,
                    KnowledgeChunk.document_id == asset["document_id"],
                )
            )
            for chunk in chunks:
                session.add(KnowledgeChunk(
                    tenant_id=tenant_id,
                    chunk_id=chunk["chunk_id"],
                    document_id=chunk["document_id"],
                    title=chunk["title"],
                    chunk_text=chunk["chunk_text"],
                    embedding=chunk["embedding"],
                    source_location=chunk["source_location"],
                    permission_scope=chunk["permission_scope"],
                    status=chunk["status"],
                ))
                seeded_chunks += 1

            await session.execute(
                delete(KnowledgeObjectLink).where(
                    KnowledgeObjectLink.tenant_id == tenant_id,
                    KnowledgeObjectLink.document_id == asset["document_id"],
                )
            )
            for object_type, object_id, object_name in asset["links"]:
                session.add(KnowledgeObjectLink(
                    tenant_id=tenant_id,
                    document_id=asset["document_id"],
                    object_type=object_type,
                    object_id=object_id,
                    object_name=object_name,
                    confidence=0.92,
                    source_location="seeded-demo-link",
                    status="committed",
                ))

            DOCUMENTS[asset["document_id"]] = document
            for chunk in chunks:
                CHUNKS[chunk["chunk_id"]] = chunk
            JOBS[f"job-{asset['document_id']}"] = {
                "job_id": f"job-{asset['document_id']}",
                "tenant_id": tenant_id,
                "asset_id": f"asset-{asset['document_id']}",
                "document_id": asset["document_id"],
                "status": "completed",
                "error": None,
                "created_at": document["updated_at"],
                "updated_at": document["updated_at"],
            }

        await session.commit()

    return {"documents": seeded_documents, "chunks": seeded_chunks}
